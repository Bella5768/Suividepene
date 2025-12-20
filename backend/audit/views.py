from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from .models import AuditLog
from .serializers import AuditLogSerializer


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet en lecture seule pour les journaux d'audit.
    Les journaux d'audit sont immuables et ne peuvent pas être modifiés.
    """
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['action', 'model_name', 'user']
    search_fields = ['object_repr', 'user__username', 'ip_address']
    ordering_fields = ['timestamp', 'action', 'model_name']
    ordering = ['-timestamp']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtre par date de début
        timestamp_after = self.request.query_params.get('timestamp_after', None)
        if timestamp_after:
            try:
                date_obj = datetime.strptime(timestamp_after, '%Y-%m-%d')
                queryset = queryset.filter(timestamp__gte=date_obj)
            except ValueError:
                pass
        
        # Filtre par date de fin
        timestamp_before = self.request.query_params.get('timestamp_before', None)
        if timestamp_before:
            try:
                date_obj = datetime.strptime(timestamp_before, '%Y-%m-%d')
                # Ajouter 23h59:59 pour inclure toute la journée
                date_obj = date_obj.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(timestamp__lte=date_obj)
            except ValueError:
                pass
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporte les journaux d'audit en Excel"""
        queryset = self.get_queryset()
        
        # Créer un workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Journaux d'Audit"
        
        # En-têtes
        headers = ['Timestamp', 'Action', 'Utilisateur', 'Modèle', 'Objet', 'IP', 'User Agent']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        action_labels = {
            'create': 'Création',
            'update': 'Modification',
            'delete': 'Suppression',
            'validate': 'Validation',
            'export': 'Export',
            'import': 'Import',
        }
        
        for row_num, log in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%d/%m/%Y %H:%M:%S'))
            ws.cell(row=row_num, column=2, value=action_labels.get(log.action, log.action))
            ws.cell(row=row_num, column=3, value=log.user.username if log.user else '-')
            ws.cell(row=row_num, column=4, value=log.model_name or '-')
            ws.cell(row=row_num, column=5, value=log.object_repr or '-')
            ws.cell(row=row_num, column=6, value=log.ip_address or '-')
            ws.cell(row=row_num, column=7, value=log.user_agent or '-')
        
        # Ajuster la largeur des colonnes
        column_widths = [20, 15, 15, 15, 40, 15, 30]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'journaux_audit_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Exporte les journaux d'audit en PDF"""
        queryset = self.get_queryset()
        
        # Créer le document PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f'journaux_audit_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        styles.add(ParagraphStyle(
            name='TitleCentered',
            parent=styles['Title'],
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        styles.add(ParagraphStyle(
            name='Heading2Left',
            parent=styles['Heading2'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold'
        ))
        
        # Fonction pour charger le logo CSIG
        def get_logo_path():
            from django.conf import settings
            import os
            project_root = settings.BASE_DIR.parent
            logo_paths = [
                str(project_root / 'logocsig.png'),
                str(project_root / 'frontend' / 'src' / 'assets' / 'logocsig.png'),
                str(settings.BASE_DIR / 'logocsig.png'),
            ]
            for path in logo_paths:
                if os.path.exists(path):
                    return path
            return None
        
        # En-tête avec logo
        logo_path = get_logo_path()
        if logo_path:
            try:
                logo = Image(logo_path, width=1.5*inch, height=1.5*inch)
                header_data = [[logo, Paragraph("<b>CSIG</b><br/>JOURNAUX D'AUDIT", styles['Title'])]]
                header_table = Table(header_data, colWidths=[2*inch, 4*inch])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(header_table)
            except Exception as e:
                print(f"Erreur lors du chargement du logo: {e}")
                title = Paragraph("CSIG - JOURNAUX D'AUDIT", styles['TitleCentered'])
                elements.append(title)
        else:
            title = Paragraph("CSIG - JOURNAUX D'AUDIT", styles['TitleCentered'])
            elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations de l'export
        info_data = [
            ['Date d\'export:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ['Nombre d\'entrées:', str(queryset.count())],
        ]
        if request.query_params.get('action'):
            info_data.append(['Action filtrée:', request.query_params.get('action')])
        if request.query_params.get('model_name'):
            info_data.append(['Modèle filtré:', request.query_params.get('model_name')])
        if request.query_params.get('timestamp_after'):
            info_data.append(['Date début:', request.query_params.get('timestamp_after')])
        if request.query_params.get('timestamp_before'):
            info_data.append(['Date fin:', request.query_params.get('timestamp_before')])
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tableau des journaux
        elements.append(Paragraph("<b>DÉTAILS DES JOURNAUX D'AUDIT</b>", styles['Heading2Left']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Préparer les données du tableau
        action_labels = {
            'create': 'Création',
            'update': 'Modification',
            'delete': 'Suppression',
            'validate': 'Validation',
            'export': 'Export',
            'import': 'Import',
        }
        
        table_data = [['Timestamp', 'Action', 'Utilisateur', 'Modèle', 'Objet', 'IP']]
        
        # Fonction pour nettoyer et encoder correctement les chaînes
        def clean_text(text, max_length=50):
            """Nettoie et encode correctement le texte pour le PDF"""
            if text is None:
                return '-'
            # Convertir en string et nettoyer
            text = str(text)
            # Remplacer les caractères problématiques
            text = text.replace('\x00', '')  # Supprimer les null bytes
            text = text.replace('\n', ' ')  # Remplacer les retours à la ligne
            text = text.replace('\r', ' ')  # Remplacer les retours chariot
            # Encoder en UTF-8 et décoder pour s'assurer que c'est valide
            try:
                text = text.encode('utf-8', errors='ignore').decode('utf-8')
            except:
                text = text.encode('ascii', errors='ignore').decode('ascii')
            # Limiter la longueur pour éviter les problèmes d'affichage
            text = text.strip()
            if len(text) > max_length:
                # Tronquer à un espace pour éviter de couper au milieu d'un mot
                truncated = text[:max_length]
                last_space = truncated.rfind(' ')
                if last_space > max_length * 0.7:  # Si on trouve un espace dans les 70% derniers caractères
                    text = truncated[:last_space] + '...'
                else:
                    text = truncated + '...'
            return text
        
        # Fonction pour nettoyer l'adresse IP
        def clean_ip(ip_address):
            """Nettoie et valide l'adresse IP"""
            if not ip_address:
                return '-'
            try:
                ip_str = str(ip_address).strip()
                # Vérifier que ce n'est pas None ou une valeur invalide
                if not ip_str or ip_str == 'None' or ip_str.lower() == 'none':
                    return '-'
                # Vérifier que ça ressemble à une IP (contient des points ou des deux-points)
                if '.' in ip_str or ':' in ip_str:
                    # Limiter la longueur (IPv6 peut être long)
                    if len(ip_str) > 45:
                        return ip_str[:42] + '...'
                    return ip_str
                else:
                    # Si ça ne ressemble pas à une IP, retourner '-'
                    return '-'
            except Exception:
                return '-'
        
        for log in queryset:
            timestamp_str = log.timestamp.strftime('%d/%m/%Y\n%H:%M:%S')
            action_str = action_labels.get(log.action, log.action)
            user_str = clean_text(log.user.username if log.user else '-', max_length=20)
            model_str = clean_text(log.model_name or '-', max_length=20)
            object_str = clean_text(log.object_repr or '-', max_length=50)
            ip_str = clean_ip(log.ip_address)
            
            table_data.append([
                timestamp_str,
                clean_text(action_str, max_length=15),
                user_str,
                model_str,
                object_str,
                ip_str
            ])
        
        # Créer le tableau
        audit_table = Table(table_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 2.5*inch, 1*inch])
        audit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Timestamp centré
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(audit_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Pied de page
        footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
        footer_style = ParagraphStyle(
            name='Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)
        
        # Construire le PDF
        doc.build(elements)
        
        return response

