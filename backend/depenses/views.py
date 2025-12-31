from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.exceptions import PermissionDenied
from django.db.models import Sum, Avg, Count, Q
from django.db import transaction, IntegrityError
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
from .models import (
    Categorie, SousCategorie, Prevision, Operation, Imputation,
    Plat, Menu, MenuPlat, FenetreCommande, RegleSubvention, Commande, CommandeLigne, Facture,
    ExtraRestauration
)
from .serializers import (
    CategorieSerializer, SousCategorieSerializer, PrevisionSerializer,
    OperationSerializer, ImputationSerializer,
    PlatSerializer, MenuSerializer, MenuPlatSerializer, FenetreCommandeSerializer,
    RegleSubventionSerializer, CommandeSerializer, CommandeLigneSerializer, CommandeCreateSerializer,
    UserSerializer, UserPermissionSerializer, ExtraRestaurationSerializer
)
from django.contrib.auth.models import User
from .filters import OperationFilter, PrevisionFilter
import pandas as pd
from django.http import HttpResponse, FileResponse
from django.conf import settings
import os
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from audit.middleware import log_audit
import csv
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags


class UserViewSet(viewsets.ModelViewSet):
    """ViewSet pour la gestion des utilisateurs et privilèges"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        # Seuls les superusers peuvent voir tous les utilisateurs
        if self.request.user.is_superuser:
            return User.objects.all().order_by('-date_joined')
        # Les autres utilisateurs ne peuvent voir que leur propre profil
        return User.objects.filter(id=self.request.user.id)
    
    def perform_create(self, serializer):
        # Seuls les superusers peuvent créer des utilisateurs
        if not self.request.user.is_superuser:
            raise PermissionDenied("Seuls les administrateurs peuvent créer des utilisateurs")
        user = serializer.save()
        # Logger l'action
        if self.request.user.is_authenticated:
            log_audit('create', self.request.user, user, metadata={'type': 'user_creation'})
    
    def perform_update(self, serializer):
        # Seuls les superusers peuvent modifier les utilisateurs
        if not self.request.user.is_superuser:
            raise PermissionDenied("Seuls les administrateurs peuvent modifier les utilisateurs")
        old_instance = self.get_object()
        user = serializer.save()
        # Logger l'action
        if self.request.user.is_authenticated:
            changes = {}
            if old_instance.is_staff != user.is_staff:
                changes['is_staff'] = {'old': old_instance.is_staff, 'new': user.is_staff}
            if old_instance.is_superuser != user.is_superuser:
                changes['is_superuser'] = {'old': old_instance.is_superuser, 'new': user.is_superuser}
            if old_instance.is_active != user.is_active:
                changes['is_active'] = {'old': old_instance.is_active, 'new': user.is_active}
            log_audit('update', self.request.user, user, changes=changes, metadata={'type': 'user_update'})
    
    def perform_destroy(self, instance):
        # Seuls les superusers peuvent supprimer des utilisateurs
        if not self.request.user.is_superuser:
            raise PermissionDenied("Seuls les administrateurs peuvent supprimer des utilisateurs")
        # Ne pas permettre la suppression de soi-même
        if instance.id == self.request.user.id:
            raise PermissionDenied("Vous ne pouvez pas supprimer votre propre compte")
        # Logger l'action
        if self.request.user.is_authenticated:
            log_audit('delete', self.request.user, instance, metadata={'type': 'user_deletion'})
        instance.delete()
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Récupère les informations de l'utilisateur actuellement connecté"""
        from .models import UserPermission
        
        # Charger les permissions depuis la base de données
        permissions = UserPermission.objects.filter(utilisateur=request.user)
        permission_serializer = UserPermissionSerializer(permissions, many=True)
        
        # Construire la réponse manuellement pour être sûr
        data = {
            'id': request.user.id,
            'username': request.user.username,
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'is_staff': request.user.is_staff,
            'is_superuser': request.user.is_superuser,
            'is_active': request.user.is_active,
            'permissions': permission_serializer.data,
        }
        
        print(f"[ME] User: {request.user.username}, Permissions: {len(permission_serializer.data)}")
        return Response(data)


class CategorieViewSet(viewsets.ModelViewSet):
    queryset = Categorie.objects.all()
    serializer_class = CategorieSerializer
    permission_classes = [IsAuthenticated]


class SousCategorieViewSet(viewsets.ModelViewSet):
    queryset = SousCategorie.objects.all()
    serializer_class = SousCategorieSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['categorie']

    def get_queryset(self):
        queryset = super().get_queryset()
        categorie_id = self.request.query_params.get('categorie', None)
        if categorie_id:
            queryset = queryset.filter(categorie_id=categorie_id)
        return queryset


class PrevisionViewSet(viewsets.ModelViewSet):
    queryset = Prevision.objects.all()
    serializer_class = PrevisionSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = PrevisionFilter

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def by_month(self, request):
        """Récupérer les prévisions pour un mois donné"""
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'error': 'Paramètre mois requis (format: YYYY-MM)'}, status=400)
        
        try:
            mois_date = datetime.strptime(mois, '%Y-%m').date()
            mois_date = mois_date.replace(day=1)
        except ValueError:
            return Response({'error': 'Format de date invalide. Utilisez YYYY-MM'}, status=400)
        
        previsions = self.queryset.filter(mois=mois_date)
        serializer = self.get_serializer(previsions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def solde(self, request, pk=None):
        """Récupérer le solde restant d'une prévision"""
        prevision = self.get_object()
        return Response({
            'montant_prevu': prevision.montant_prevu,
            'montant_impute': prevision.montant_impute,
            'solde_restant': prevision.solde_restant
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporter les prévisions en CSV"""
        previsions = self.queryset.all()
        
        mois = request.query_params.get('mois')
        if mois:
            previsions = previsions.filter(mois__startswith=mois)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="previsions.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'ID', 'Mois', 'Catégorie', 'Code Catégorie', 'Sous-Catégorie',
            'Montant Prévu', 'Statut', 'Montant Imputé', 'Solde Restant',
            'Créé par', 'Créé le'
        ])
        
        for prev in previsions:
            writer.writerow([
                prev.id, prev.mois, prev.categorie.nom, prev.categorie.code,
                prev.sous_categorie.nom if prev.sous_categorie else '',
                prev.montant_prevu, prev.get_statut_display(),
                prev.montant_impute, prev.solde_restant,
                prev.created_by.username if prev.created_by else '',
                prev.created_at
            ])
        
        if request.user.is_authenticated:
            log_audit('export', request.user, None, metadata={'type': 'previsions_csv'})
        
        return response

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Importer des prévisions depuis un fichier CSV"""
        if 'file' not in request.FILES:
            return Response({'error': 'Fichier CSV requis'}, status=400)
        
        file = request.FILES['file']
        
        try:
            df = pd.read_csv(file, delimiter=';', encoding='utf-8-sig')
            
            required_columns = ['Mois', 'Catégorie', 'Montant Prévu']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response({
                    'error': f'Colonnes manquantes: {", ".join(missing_columns)}'
                }, status=400)
            
            imported = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    categorie_code = row.get('Code Catégorie') or row.get('Catégorie')
                    categorie, _ = Categorie.objects.get_or_create(
                        code=categorie_code[:20],
                        defaults={'nom': row.get('Catégorie', categorie_code)}
                    )
                    
                    sous_categorie = None
                    if pd.notna(row.get('Sous-Catégorie')):
                        sous_categorie, _ = SousCategorie.objects.get_or_create(
                            categorie=categorie,
                            nom=row['Sous-Catégorie'][:100]
                        )
                    
                    mois = pd.to_datetime(row['Mois']).date().replace(day=1)
                    
                    prevision, created = Prevision.objects.get_or_create(
                        mois=mois,
                        categorie=categorie,
                        sous_categorie=sous_categorie,
                        defaults={
                            'montant_prevu': float(row['Montant Prévu']),
                            'statut': row.get('Statut', 'draft'),
                            'created_by': request.user if request.user.is_authenticated else None
                        }
                    )
                    
                    if not created:
                        prevision.montant_prevu = float(row['Montant Prévu'])
                        prevision.save()
                    
                    imported += 1
                    
                    if request.user.is_authenticated:
                        log_audit('import', request.user, prevision, metadata={'source': 'csv'})
                        
                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            return Response({
                'imported': imported,
                'errors': errors[:10],
                'total_errors': len(errors)
            }, status=201 if imported > 0 else 400)
            
        except Exception as e:
            return Response({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=400)


class OperationViewSet(viewsets.ModelViewSet):
    queryset = Operation.objects.all()
    serializer_class = OperationSerializer
    permission_classes = [IsAuthenticated]
    filterset_class = OperationFilter

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def by_date_range(self, request):
        """Récupérer les opérations dans une plage de dates"""
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        
        if not date_debut or not date_fin:
            return Response({'error': 'date_debut et date_fin requis (format: YYYY-MM-DD)'}, status=400)
        
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide. Utilisez YYYY-MM-DD'}, status=400)
        
        operations = self.queryset.filter(
            date_operation__gte=date_debut,
            date_operation__lte=date_fin
        )
        serializer = self.get_serializer(operations, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def totals_by_day(self, request):
        """Totaux par jour dans une plage de dates"""
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        
        if not date_debut or not date_fin:
            return Response({'error': 'date_debut et date_fin requis'}, status=400)
        
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        totals = self.queryset.filter(
            date_operation__gte=date_debut,
            date_operation__lte=date_fin
        ).values('date_operation').annotate(
            total=Sum('montant_depense'),
            count=Count('id')
        ).order_by('date_operation')
        
        return Response(list(totals))

    @action(detail=False, methods=['get'])
    def totals_by_week(self, request):
        """Totaux par semaine dans une plage de dates"""
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        
        if not date_debut or not date_fin:
            return Response({'error': 'date_debut et date_fin requis'}, status=400)
        
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        totals = self.queryset.filter(
            date_operation__gte=date_debut,
            date_operation__lte=date_fin
        ).values('semaine_iso').annotate(
            total=Sum('montant_depense'),
            count=Count('id')
        ).order_by('semaine_iso')
        
        return Response(list(totals))

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Exporter les opérations en CSV"""
        operations = self.queryset.all()
        
        # Appliquer les filtres si présents
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        categorie_id = request.query_params.get('categorie')
        
        if date_debut:
            operations = operations.filter(date_operation__gte=date_debut)
        if date_fin:
            operations = operations.filter(date_operation__lte=date_fin)
        if categorie_id:
            operations = operations.filter(categorie_id=categorie_id)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="operations.csv"'
        response.write('\ufeff')  # BOM pour Excel UTF-8
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'ID', 'Date Opération', 'Jour', 'Semaine ISO', 'Catégorie', 'Code Catégorie',
            'Sous-Catégorie', 'Unités', 'Prix Unitaire', 'Montant Dépensé',
            'Description', 'Créé par', 'Créé le'
        ])
        
        for op in operations:
            writer.writerow([
                op.id, op.date_operation, op.jour, op.semaine_iso,
                op.categorie.nom, op.categorie.code,
                op.sous_categorie.nom if op.sous_categorie else '',
                op.unites, op.prix_unitaire, op.montant_depense,
                op.description, op.created_by.username if op.created_by else '',
                op.created_at
            ])
        
        if request.user.is_authenticated:
            log_audit('export', request.user, None, metadata={'type': 'operations_csv'})
        
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporter les opérations en Excel"""
        operations = self.get_queryset()
        
        # Appliquer les filtres si présents
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        categorie_id = request.query_params.get('categorie')
        
        if date_debut:
            operations = operations.filter(date_operation__gte=date_debut)
        if date_fin:
            operations = operations.filter(date_operation__lte=date_fin)
        if categorie_id:
            operations = operations.filter(categorie_id=categorie_id)
        
        # Optimiser les requêtes
        operations = operations.select_related('categorie', 'sous_categorie', 'created_by')
        
        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Opérations"
        
        # Styles
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # En-têtes
        headers = [
            'ID', 'Date Opération', 'Jour', 'Semaine ISO', 'Catégorie', 'Code Catégorie',
            'Sous-Catégorie', 'Unités', 'Prix Unitaire (GNF)', 'Montant Dépensé (GNF)',
            'Description', 'Créé par', 'Créé le'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        for row_idx, op in enumerate(operations, start=2):
            ws.cell(row=row_idx, column=1, value=op.id)
            ws.cell(row=row_idx, column=2, value=op.date_operation.strftime('%Y-%m-%d') if op.date_operation else '')
            ws.cell(row=row_idx, column=3, value=op.jour)
            ws.cell(row=row_idx, column=4, value=op.semaine_iso)
            ws.cell(row=row_idx, column=5, value=op.categorie.nom if op.categorie else '')
            ws.cell(row=row_idx, column=6, value=op.categorie.code if op.categorie else '')
            ws.cell(row=row_idx, column=7, value=op.sous_categorie.nom if op.sous_categorie else '')
            ws.cell(row=row_idx, column=8, value=float(op.unites))
            ws.cell(row=row_idx, column=9, value=float(op.prix_unitaire))
            ws.cell(row=row_idx, column=10, value=float(op.montant_depense))
            ws.cell(row=row_idx, column=11, value=op.description or '')
            ws.cell(row=row_idx, column=12, value=op.created_by.username if op.created_by else '')
            ws.cell(row=row_idx, column=13, value=op.created_at.strftime('%Y-%m-%d %H:%M:%S') if op.created_at else '')
        
        # Ajuster la largeur des colonnes
        column_widths = [8, 15, 10, 12, 25, 12, 20, 10, 18, 18, 30, 15, 20]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'operations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        if request.user.is_authenticated:
            log_audit('export', request.user, None, metadata={'type': 'operations_excel'})
        
        return response
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Exporter les opérations en PDF"""
        operations = self.get_queryset()
        
        # Appliquer les filtres si présents
        date_debut = request.query_params.get('date_debut')
        date_fin = request.query_params.get('date_fin')
        categorie_id = request.query_params.get('categorie')
        
        if date_debut:
            operations = operations.filter(date_operation__gte=date_debut)
        if date_fin:
            operations = operations.filter(date_operation__lte=date_fin)
        if categorie_id:
            operations = operations.filter(categorie_id=categorie_id)
        
        # Optimiser les requêtes
        operations = operations.select_related('categorie', 'sous_categorie', 'created_by').order_by('-date_operation')
        
        # Créer le document PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f'operations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles personnalisés
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        styles.add(ParagraphStyle(
            name='TitleCentered',
            parent=styles['Title'],
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        styles.add(ParagraphStyle(
            name='Heading2Left',
            parent=styles['Heading2'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            fontName='Helvetica-Bold'
        ))
        
        # Fonction pour charger le logo CSIG
        def get_logo_path():
            project_root = settings.BASE_DIR.parent
            logo_paths = [
                str(project_root / 'logocsig.png'),
                str(project_root / 'frontend' / 'src' / 'assets' / 'logocsig.png'),
                str(settings.BASE_DIR / 'logocsig.png'),
            ]
            for path in logo_paths:
                if os.path.exists(path):
                    return path
            return None
        
        # En-tête avec logo
        logo_path = get_logo_path()
        if logo_path:
            try:
                logo = Image(logo_path, width=1.5*inch, height=1.5*inch)
                header_data = [[logo, Paragraph("<b>CSIG</b><br/>OPÉRATIONS", styles['Title'])]]
                header_table = Table(header_data, colWidths=[2*inch, 4*inch])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(header_table)
            except Exception as e:
                print(f"Erreur lors du chargement du logo: {e}")
                title = Paragraph("CSIG - OPÉRATIONS", styles['TitleCentered'])
                elements.append(title)
        else:
            title = Paragraph("CSIG - OPÉRATIONS", styles['TitleCentered'])
            elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations de l'export
        info_data = [
            ['Date d\'export:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ['Nombre d\'opérations:', str(operations.count())],
        ]
        if date_debut:
            info_data.append(['Date début:', date_debut])
        if date_fin:
            info_data.append(['Date fin:', date_fin])
        if categorie_id:
            try:
                from .models import Categorie
                categorie = Categorie.objects.get(pk=categorie_id)
                info_data.append(['Catégorie:', categorie.nom])
            except:
                pass
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tableau des opérations
        elements.append(Paragraph("<b>DÉTAILS DES OPÉRATIONS</b>", styles['Heading2Left']))
        elements.append(Spacer(1, 0.1*inch))
        
        # Fonction pour nettoyer le texte
        def clean_text(text, max_length=50):
            if text is None:
                return '-'
            text = str(text)
            # Remplacer les caractères problématiques
            text = text.replace('\x00', '')  # Supprimer les null bytes
            text = text.replace('\n', ' ')  # Remplacer les retours à la ligne
            text = text.replace('\r', ' ')  # Remplacer les retours chariot
            # Encoder en UTF-8
            try:
                text = text.encode('utf-8', errors='ignore').decode('utf-8')
            except:
                text = text.encode('ascii', errors='ignore').decode('ascii')
            # Limiter la longueur
            text = text.strip()
            if len(text) > max_length:
                # Tronquer à un espace pour éviter de couper au milieu d'un mot
                truncated = text[:max_length]
                last_space = truncated.rfind(' ')
                if last_space > max_length * 0.7:
                    text = truncated[:last_space] + '...'
                else:
                    text = truncated + '...'
            return text
        
        # Fonction pour formater les montants
        def format_montant(montant):
            """Formate un montant Decimal en string propre"""
            try:
                if montant is None:
                    return '-'
                # Convertir en Decimal si nécessaire
                if isinstance(montant, (int, float)):
                    montant = Decimal(str(montant))
                # Formater avec séparateur de milliers
                return f"{montant:,.0f} GNF"
            except Exception as e:
                print(f"Erreur formatage montant: {e}, valeur: {montant}")
                return str(montant) if montant else '-'
        
        # Préparer les données du tableau
        table_data = [['Date', 'Catégorie', 'Sous-Cat.', 'Unités', 'Prix Unit.', 'Montant', 'Description']]
        
        for op in operations:
            table_data.append([
                op.date_operation.strftime('%d/%m/%Y') if op.date_operation else '-',
                clean_text(op.categorie.nom if op.categorie else '-', max_length=25),
                clean_text(op.sous_categorie.nom if op.sous_categorie else '-', max_length=20),
                f"{float(op.unites):,.2f}",
                format_montant(op.prix_unitaire),
                format_montant(op.montant_depense),
                clean_text(op.description or '-', max_length=40)
            ])
        
        # Créer le tableau avec largeurs ajustées
        audit_table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1*inch, 0.8*inch, 1.1*inch, 1.1*inch, 1.5*inch])
        audit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (5, -1), 'RIGHT'),  # Colonnes numériques alignées à droite (Unités, Prix, Montant)
            ('ALIGN', (6, 0), (6, -1), 'LEFT'),  # Description alignée à gauche
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('WORDWRAP', (0, 0), (-1, -1)),  # Permettre le retour à la ligne
        ]))
        
        elements.append(audit_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Totaux
        if operations.exists():
            total_montant = sum(op.montant_depense for op in operations)
            total_data = [
                ['Total des opérations:', str(operations.count())],
                ['Total montant dépensé:', f"{total_montant:,.0f} GNF"],
            ]
            total_table = Table(total_data, colWidths=[3*inch, 3*inch])
            total_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
            ]))
            elements.append(total_table)
        
        # Pied de page
        elements.append(Spacer(1, 0.3*inch))
        footer_text = f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
        footer_style = ParagraphStyle(
            name='Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            fontName='Helvetica'
        )
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)
        
        # Construire le PDF
        doc.build(elements)
        
        if request.user.is_authenticated:
            log_audit('export', request.user, None, metadata={'type': 'operations_pdf'})
        
        return response

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Importer des opérations depuis un fichier CSV"""
        if 'file' not in request.FILES:
            return Response({'error': 'Fichier CSV requis'}, status=400)
        
        file = request.FILES['file']
        
        try:
            # Lire le CSV
            df = pd.read_csv(file, delimiter=';', encoding='utf-8-sig')
            
            # Vérifier les colonnes requises
            required_columns = ['Date Opération', 'Catégorie', 'Unités', 'Prix Unitaire', 'Montant Dépensé']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response({
                    'error': f'Colonnes manquantes: {", ".join(missing_columns)}'
                }, status=400)
            
            imported = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Trouver ou créer la catégorie
                    categorie_code = row.get('Code Catégorie') or row.get('Catégorie')
                    categorie, _ = Categorie.objects.get_or_create(
                        code=categorie_code[:20],
                        defaults={'nom': row.get('Catégorie', categorie_code)}
                    )
                    
                    # Trouver ou créer la sous-catégorie si présente
                    sous_categorie = None
                    if pd.notna(row.get('Sous-Catégorie')):
                        sous_categorie, _ = SousCategorie.objects.get_or_create(
                            categorie=categorie,
                            nom=row['Sous-Catégorie'][:100]
                        )
                    
                    # Créer l'opération
                    operation = Operation.objects.create(
                        date_operation=pd.to_datetime(row['Date Opération']).date(),
                        categorie=categorie,
                        sous_categorie=sous_categorie,
                        unites=float(row['Unités']),
                        prix_unitaire=float(row['Prix Unitaire']),
                        description=row.get('Description', '')[:500],
                        created_by=request.user if request.user.is_authenticated else None
                    )
                    
                    imported += 1
                    
                    if request.user.is_authenticated:
                        log_audit('import', request.user, operation, metadata={'source': 'csv'})
                        
                except Exception as e:
                    errors.append(f"Ligne {index + 2}: {str(e)}")
            
            return Response({
                'imported': imported,
                'errors': errors[:10],  # Limiter à 10 erreurs
                'total_errors': len(errors)
            }, status=201 if imported > 0 else 400)
            
        except Exception as e:
            return Response({'error': f'Erreur lors de la lecture du fichier: {str(e)}'}, status=400)


class ImputationViewSet(viewsets.ModelViewSet):
    queryset = Imputation.objects.all()
    serializer_class = ImputationSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def multi_impute(self, request):
        """Multi-imputation : imputer une opération sur plusieurs prévisions"""
        operation_id = request.data.get('operation_id')
        imputations = request.data.get('imputations', [])
        
        if not operation_id or not imputations:
            return Response({'error': 'operation_id et imputations requis'}, status=400)
        
        try:
            operation = Operation.objects.get(pk=operation_id)
        except Operation.DoesNotExist:
            return Response({'error': 'Opération non trouvée'}, status=404)
        
        total_impute = sum(imp.get('montant_impute', 0) for imp in imputations)
        if total_impute > operation.montant_depense:
            return Response({
                'error': f'Le total imputé ({total_impute} GNF) dépasse le montant de l\'opération ({operation.montant_depense} GNF)'
            }, status=400)
        
        created_imputations = []
        errors = []
        
        with transaction.atomic():
            for imp_data in imputations:
                prevision_id = imp_data.get('prevision_id')
                montant = imp_data.get('montant_impute')
                
                try:
                    prevision = Prevision.objects.get(pk=prevision_id)
                    if montant > prevision.solde_restant:
                        errors.append(f'Prévision {prevision_id}: solde insuffisant')
                        continue
                    
                    imputation, created = Imputation.objects.get_or_create(
                        operation=operation,
                        prevision=prevision,
                        defaults={
                            'montant_impute': montant,
                            'created_by': request.user
                        }
                    )
                    if not created:
                        imputation.montant_impute = montant
                        imputation.save()
                    
                    created_imputations.append(imputation)
                except Prevision.DoesNotExist:
                    errors.append(f'Prévision {prevision_id} non trouvée')
        
        if errors:
            return Response({'errors': errors, 'created': len(created_imputations)}, status=207)
        
        serializer = ImputationSerializer(created_imputations, many=True)
        return Response(serializer.data, status=201)


class RapportViewSet(viewsets.ViewSet):
    from rest_framework.authentication import SessionAuthentication, BasicAuthentication
    from rest_framework_simplejwt.authentication import JWTAuthentication
    
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def mensuel(self, request):
        """Générer un rapport mensuel avec totaux, écarts et moyenne journalière"""
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'error': 'Paramètre mois requis (format: YYYY-MM)'}, status=400)
        
        try:
            mois_date = datetime.strptime(mois, '%Y-%m').date()
            mois_date = mois_date.replace(day=1)
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        # Calculer le dernier jour du mois
        dernier_jour = calendar.monthrange(mois_date.year, mois_date.month)[1]
        date_fin = mois_date.replace(day=dernier_jour)
        
        # Opérations du mois
        operations = Operation.objects.filter(
            date_operation__gte=mois_date,
            date_operation__lte=date_fin
        )
        
        # Prévisions du mois
        previsions = Prevision.objects.filter(mois=mois_date)
        
        # Totaux par catégorie
        totals_operations = operations.values('categorie__code', 'categorie__nom').annotate(
            total=Sum('montant_depense'),
            count=Count('id')
        )
        
        # Calcul des écarts
        rapport_data = {
            'mois': mois,
            'date_debut': mois_date,
            'date_fin': date_fin,
            'total_depenses': operations.aggregate(Sum('montant_depense'))['montant_depense__sum'] or 0,
            'total_prevu': previsions.aggregate(Sum('montant_prevu'))['montant_prevu__sum'] or 0,
            'nombre_jours': dernier_jour,
            'nombre_operations': operations.count(),
            'moyenne_journaliere': 0,
            'categories': []
        }
        
        # Calcul de la moyenne journalière
        jours_avec_operations = operations.values('date_operation').distinct().count()
        if jours_avec_operations > 0:
            rapport_data['moyenne_journaliere'] = rapport_data['total_depenses'] / jours_avec_operations
        
        # Détails par catégorie
        for tot in totals_operations:
            categorie_code = tot['categorie__code']
            prevision_cat = previsions.filter(categorie__code=categorie_code).first()
            
            categorie_data = {
                'categorie_code': categorie_code,
                'categorie_nom': tot['categorie__nom'],
                'total_depense': tot['total'],
                'nombre_operations': tot['count'],
                'montant_prevu': prevision_cat.montant_prevu if prevision_cat else 0,
                'ecart': 0
            }
            
            if prevision_cat:
                categorie_data['ecart'] = tot['total'] - prevision_cat.montant_prevu
            
            rapport_data['categories'].append(categorie_data)
        
        # Écart global
        rapport_data['ecart_global'] = rapport_data['total_depenses'] - rapport_data['total_prevu']
        
        return Response(rapport_data)

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Exporter le rapport mensuel en PDF"""
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'error': 'Paramètre mois requis'}, status=400)
        
        # Générer le rapport
        rapport_response = self.mensuel(request)
        if rapport_response.status_code != 200:
            return rapport_response
        
        rapport_data = rapport_response.data
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_{mois}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title = Paragraph(f"Rapport Mensuel - {mois}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Informations générales
        info_data = [
            ['Période', f"{rapport_data['date_debut']} au {rapport_data['date_fin']}"],
            ['Total Dépenses', f"{rapport_data['total_depenses']:,.2f} GNF"],
            ['Total Prévu', f"{rapport_data['total_prevu']:,.2f} GNF"],
            ['Écart Global', f"{rapport_data['ecart_global']:,.2f} GNF"],
            ['Nombre d\'opérations', str(rapport_data['nombre_operations'])],
            ['Moyenne Journalière', f"{rapport_data['moyenne_journaliere']:,.2f} GNF"],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tableau des catégories
        cat_headers = ['Catégorie', 'Total Dépense', 'Montant Prévu', 'Écart', 'Nb Opérations']
        cat_data = [cat_headers]
        
        for cat in rapport_data['categories']:
            cat_data.append([
                f"{cat['categorie_code']} - {cat['categorie_nom']}",
                f"{cat['total_depense']:,.2f} GNF",
                f"{cat['montant_prevu']:,.2f} GNF",
                f"{cat['ecart']:,.2f} GNF",
                str(cat['nombre_operations'])
            ])
        
        cat_table = Table(cat_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(cat_table)
        
        doc.build(elements)
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exporter le rapport mensuel en Excel"""
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'error': 'Paramètre mois requis'}, status=400)
        
        # Générer le rapport
        rapport_response = self.mensuel(request)
        if rapport_response.status_code != 200:
            return rapport_response
        
        rapport_data = rapport_response.data
        
        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Rapport {mois}"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        # Titre
        ws['A1'] = f"Rapport Mensuel - {mois}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        # Informations générales
        row = 3
        info_labels = ['Période', 'Total Dépenses', 'Total Prévu', 'Écart Global', 'Nombre d\'opérations', 'Moyenne Journalière']
        info_values = [
            f"{rapport_data['date_debut']} au {rapport_data['date_fin']}",
            f"{rapport_data['total_depenses']:,.2f} GNF",
            f"{rapport_data['total_prevu']:,.2f} GNF",
            f"{rapport_data['ecart_global']:,.2f} GNF",
            str(rapport_data['nombre_operations']),
            f"{rapport_data['moyenne_journaliere']:,.2f} GNF"
        ]
        
        for label, value in zip(info_labels, info_values):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Tableau des catégories
        row += 2
        headers = ['Catégorie', 'Total Dépense', 'Montant Prévu', 'Écart', 'Nb Opérations']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        for cat in rapport_data['categories']:
            ws.cell(row=row, column=1, value=f"{cat['categorie_code']} - {cat['categorie_nom']}")
            ws.cell(row=row, column=2, value=f"{cat['total_depense']:,.2f} GNF")
            ws.cell(row=row, column=3, value=f"{cat['montant_prevu']:,.2f} GNF")
            ws.cell(row=row, column=4, value=f"{cat['ecart']:,.2f} GNF")
            ws.cell(row=row, column=5, value=cat['nombre_operations'])
            row += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="rapport_{mois}.xlsx"'
        wb.save(response)
        return response


# ============================================================================
# VIEWSETS RESTAURATION / CANTINE
# ============================================================================

class PlatViewSet(viewsets.ModelViewSet):
    """Gestion des plats (CRUD)"""
    queryset = Plat.objects.all()
    serializer_class = PlatSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['categorie_restau', 'actif']
    search_fields = ['nom', 'description']
    ordering = ['categorie_restau', 'nom']


class MenuViewSet(viewsets.ModelViewSet):
    """Gestion des menus du jour"""
    queryset = Menu.objects.all()
    serializer_class = MenuSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['date_menu']
    ordering = ['-date_menu']
    
    def get_serializer_context(self):
        """Ajouter la requête au contexte pour générer les liens absolus"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def create(self, request, *args, **kwargs):
        """Créer un menu et ajouter automatiquement tous les plats actifs"""
        response = super().create(request, *args, **kwargs)
        
        if response.status_code == 201:
            # Récupérer le menu créé
            menu_id = response.data.get('id')
            menu = Menu.objects.get(pk=menu_id)
            
            # Ajouter tous les plats actifs au menu
            plats_actifs = Plat.objects.filter(actif=True)
            ordre = 0
            for plat in plats_actifs:
                MenuPlat.objects.create(
                    menu=menu,
                    plat=plat,
                    prix_jour=plat.prix_standard,
                    ordre=ordre
                )
                ordre += 1
            
            # Recharger le serializer avec les plats
            serializer = self.get_serializer(menu)
            return Response(serializer.data, status=201)
        
        return response
    
    def get_queryset(self):
        """Filtrer par date_menu si fourni"""
        queryset = super().get_queryset()
        date_menu = self.request.query_params.get('date_menu')
        if date_menu:
            try:
                date_obj = datetime.strptime(date_menu, '%Y-%m-%d').date()
                queryset = queryset.filter(date_menu=date_obj)
            except ValueError:
                pass
        return queryset
    
    @action(detail=False, methods=['get'])
    def by_date_range(self, request):
        """Récupérer les menus dans une plage de dates"""
        date_debut = request.query_params.get('from')
        date_fin = request.query_params.get('to')
        
        if not date_debut or not date_fin:
            return Response({'error': 'Paramètres from et to requis (format: YYYY-MM-DD)'}, status=400)
        
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Format de date invalide. Utilisez YYYY-MM-DD'}, status=400)
        
        menus = self.queryset.filter(date_menu__gte=date_debut, date_menu__lte=date_fin)
        serializer = self.get_serializer(menus, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def publier(self, request, pk=None):
        """Publier un menu"""
        import secrets
        menu = self.get_object()
        menu.publication_at = timezone.now()
        # Générer un token si pas déjà présent
        if not menu.token_public:
            menu.token_public = secrets.token_urlsafe(32)
        menu.save()
        
        if request.user.is_authenticated:
            log_audit('update', request.user, menu, metadata={'action': 'publication_menu'})
        
        serializer = self.get_serializer(menu)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def ajouter_plat(self, request, pk=None):
        """Ajouter un plat au menu (seulement si le plat est actif)"""
        menu = self.get_object()
        plat_id = request.data.get('plat_id')
        prix_jour = request.data.get('prix_jour')
        stock_max = request.data.get('stock_max')
        ordre = request.data.get('ordre', 0)
        
        if not plat_id or not prix_jour:
            return Response({'error': 'plat_id et prix_jour requis'}, status=400)
        
        try:
            plat = Plat.objects.get(pk=plat_id, actif=True)  # Vérifier que le plat est actif
            menu_plat, created = MenuPlat.objects.get_or_create(
                menu=menu,
                plat=plat,
                defaults={
                    'prix_jour': prix_jour,
                    'stock_max': stock_max,
                    'ordre': ordre
                }
            )
            if not created:
                menu_plat.prix_jour = prix_jour
                if stock_max is not None:
                    menu_plat.stock_max = stock_max
                menu_plat.ordre = ordre
                menu_plat.save()
            
            serializer = MenuPlatSerializer(menu_plat)
            return Response(serializer.data, status=201 if created else 200)
        except Plat.DoesNotExist:
            return Response({'error': 'Plat non trouvé ou non disponible (inactif)'}, status=404)
    
    @action(detail=True, methods=['delete'])
    def retirer_plat(self, request, pk=None):
        """Retirer un plat du menu"""
        menu = self.get_object()
        menu_plat_id = request.data.get('menu_plat_id')
        
        if not menu_plat_id:
            return Response({'error': 'menu_plat_id requis'}, status=400)
        
        try:
            menu_plat = MenuPlat.objects.get(pk=menu_plat_id, menu=menu)
            menu_plat.delete()
            return Response({'message': 'Plat retiré du menu'}, status=200)
        except MenuPlat.DoesNotExist:
            return Response({'error': 'Plat non trouvé dans ce menu'}, status=404)


class MenuPlatViewSet(viewsets.ModelViewSet):
    """Gestion des plats d'un menu"""
    queryset = MenuPlat.objects.all()
    serializer_class = MenuPlatSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['menu', 'plat']


class FenetreCommandeViewSet(viewsets.ModelViewSet):
    """Gestion des fenêtres de commande"""
    queryset = FenetreCommande.objects.all()
    serializer_class = FenetreCommandeSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['actif', 'categorie_restau']
    ordering = ['categorie_restau']


class RegleSubventionViewSet(viewsets.ModelViewSet):
    """Gestion des règles de subvention"""
    queryset = RegleSubvention.objects.all()
    serializer_class = RegleSubventionSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['actif', 'type_subvention']
    ordering = ['-effectif_de', '-created_at']
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Récupérer la règle active pour une date donnée"""
        date_str = request.query_params.get('date')
        if date_str:
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Format de date invalide'}, status=400)
        else:
            date = timezone.now().date()
        
        regle = RegleSubvention.objects.filter(
            actif=True,
            effectif_de__lte=date,
            effectif_a__gte=date
        ).first()
        
        if regle:
            serializer = self.get_serializer(regle)
            return Response(serializer.data)
        return Response({'message': 'Aucune règle active pour cette date'}, status=404)


class CommandeViewSet(viewsets.ModelViewSet):
    """Gestion des commandes"""
    queryset = Commande.objects.all()
    serializer_class = CommandeSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['date_commande', 'etat', 'utilisateur']
    ordering = ['-date_commande', '-created_at']
    
    def perform_update(self, serializer):
        """Envoyer un email de confirmation quand le gestionnaire valide la commande"""
        ancien_etat = serializer.instance.etat
        commande = serializer.save()
        nouveau_etat = commande.etat
        
        # Si la commande passe de brouillon à validée, envoyer l'email
        if ancien_etat == 'brouillon' and nouveau_etat == 'validee':
            try:
                email_result = envoyer_email_confirmation(commande)
                if email_result and email_result.get('success'):
                    print(f"[OK] Email de confirmation envoye a {email_result.get('email')}")
            except Exception as e:
                print(f"[ERREUR] Erreur lors de l'envoi de l'email: {e}")
    
    def get_queryset(self):
        """Filtrer par utilisateur si non-admin et sans permission de validation"""
        queryset = super().get_queryset()
        
        # Les superusers et staff voient tout
        if self.request.user.is_staff or self.request.user.is_superuser:
            return queryset
        
        # Vérifier si l'utilisateur a la permission de valider les commandes
        from .models import UserPermission
        has_validation_permission = UserPermission.objects.filter(
            utilisateur=self.request.user,
            fonctionnalite='restauration_valider_commandes',
            peut_modifier=True
        ).exists()
        
        # Si l'utilisateur a la permission de validation, il peut voir toutes les commandes
        if has_validation_permission:
            return queryset
        
        # Sinon, il ne voit que ses propres commandes
        return queryset.filter(utilisateur=self.request.user)
    
    def perform_create(self, serializer):
        """Créer une commande pour l'utilisateur connecté"""
        try:
            serializer.save(utilisateur=self.request.user)
        except IntegrityError as e:
            # Si une commande existe déjà pour cet utilisateur et cette date, la mettre à jour
            if 'unique' in str(e).lower() or 'utilisateur_id' in str(e).lower():
                data = serializer.validated_data
                commande = Commande.objects.get(
                    utilisateur=self.request.user,
                    date_commande=data.get('date_commande')
                )
                # Mettre à jour la commande existante
                for field, value in data.items():
                    setattr(commande, field, value)
                commande.save()
                serializer.instance = commande
            else:
                raise
    
    @action(detail=False, methods=['post'])
    def creer_avec_lignes(self, request):
        """Créer une commande avec ses lignes en une seule requête"""
        create_serializer = CommandeCreateSerializer(data=request.data)
        if not create_serializer.is_valid():
            return Response(create_serializer.errors, status=400)
        
        date_commande = create_serializer.validated_data['date_commande']
        lignes_data = create_serializer.validated_data['lignes']
        
        # Limiter à un seul plat par commande
        if len(lignes_data) > 1:
            return Response({'error': 'Vous ne pouvez commander qu\'un seul plat par commande'}, status=400)
        
        # Vérifier que la quantité totale est de 1
        quantite_totale = sum(int(ligne_data['quantite']) for ligne_data in lignes_data)
        if quantite_totale > 1:
            return Response({'error': 'Vous ne pouvez commander qu\'un seul plat (quantité = 1)'}, status=400)
        
        # Vérifier les fenêtres de commande pour chaque plat
        # Les admins/staff peuvent commander sans restriction horaire
        # Les utilisateurs normaux sont limités à 13h00 GMT
        erreurs_fenetre = []
        
        # Vérifier si l'utilisateur est authentifié et a des droits d'admin
        est_admin_ou_staff = request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)
        
        # Si ce n'est pas un admin/staff, vérifier les fenêtres de commande
        if not est_admin_ou_staff:
            for ligne_data in lignes_data:
                menu_plat_id = ligne_data['menu_plat_id']
                try:
                    menu_plat = MenuPlat.objects.select_related('plat').get(pk=menu_plat_id)
                    categorie_restau = menu_plat.plat.categorie_restau
                    
                    # Vérifier la fenêtre de commande (13h00 GMT pour les utilisateurs normaux)
                    if not FenetreCommande.est_dans_fenetre(categorie_restau, date_commande, est_public=False):
                        fenetre = FenetreCommande.objects.filter(categorie_restau=categorie_restau, actif=True).first()
                        if fenetre:
                            erreurs_fenetre.append(
                                f"Fenêtre de commande fermée pour {menu_plat.plat.nom} "
                                f"({menu_plat.plat.get_categorie_restau_display()}). "
                                f"Heure limite: {fenetre.heure_limite.strftime('%H:%M')}"
                            )
                        else:
                            # Si pas de fenêtre configurée, appliquer la limite par défaut de 13h00 GMT
                            from datetime import time
                            heure_limite_public = time(18, 0, 0)
                            heure_actuelle = timezone.now().time()
                            if date_commande == timezone.now().date() and heure_actuelle > heure_limite_public:
                                erreurs_fenetre.append(
                                    f"Fenêtre de commande fermée pour {menu_plat.plat.nom}. "
                                    f"Heure limite: 18:00 GMT"
                                )
                except MenuPlat.DoesNotExist:
                    pass
            
            if erreurs_fenetre:
                return Response({'error': 'Fenêtre de commande fermée', 'details': erreurs_fenetre}, status=400)
        
        try:
            with transaction.atomic():
                # Vérifier si une commande existe déjà pour cet utilisateur et cette date
                commande, created = Commande.objects.get_or_create(
                    utilisateur=request.user,
                    date_commande=date_commande,
                    defaults={'etat': 'brouillon'}
                )
                
                # Si la commande existait déjà, supprimer les anciennes lignes et réinitialiser l'état
                if not created:
                    commande.lignes.all().delete()
                    commande.etat = 'brouillon'
                    commande.save()
                
                # Créer les lignes
                for ligne_data in lignes_data:
                    menu_plat_id = ligne_data['menu_plat_id']
                    quantite = int(ligne_data['quantite'])
                    
                    try:
                        menu_plat = MenuPlat.objects.get(pk=menu_plat_id)
                        
                        # Vérifier le stock
                        stock_restant = menu_plat.get_stock_restant()
                        if stock_restant is not None and quantite > stock_restant:
                            raise ValueError(f"Stock insuffisant pour {menu_plat.plat.nom}. Stock restant: {stock_restant}")
                        
                        CommandeLigne.objects.create(
                            commande=commande,
                            menu_plat=menu_plat,
                            quantite=quantite,
                            prix_unitaire=menu_plat.prix_jour
                        )
                    except MenuPlat.DoesNotExist:
                        raise ValueError(f"MenuPlat {menu_plat_id} non trouvé")
                
                # Calculer les montants
                commande.calculer_montants()
                commande.save()
                
                if request.user.is_authenticated:
                    log_audit('create', request.user, commande, metadata={'type': 'commande_restauration'})
                
                serializer = self.get_serializer(commande)
                return Response(serializer.data, status=201)
        
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'Erreur lors de la création: {str(e)}'}, status=500)
    
    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valider une commande et créer l'opération budgétaire"""
        commande = self.get_object()
        
        # Vérifier les permissions de validation
        from .models import UserPermission
        
        # Les superusers et staff ont toujours la permission
        est_admin_ou_staff = request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser)
        
        if est_admin_ou_staff:
            has_permission = True
        else:
            # Vérifier si l'utilisateur a la permission de valider les commandes
            permission = UserPermission.objects.filter(
                utilisateur=request.user,
                fonctionnalite='restauration_valider_commandes',
                peut_modifier=True
            ).first()
            has_permission = permission is not None
        
        if not has_permission:
            raise PermissionDenied("Vous n'avez pas la permission de valider les commandes")
        
        if commande.etat != 'brouillon':
            return Response({'error': 'Seules les commandes en brouillon peuvent être validées'}, status=400)
        
        # Vérifier la fenêtre de commande avant de valider
        # Les admins/staff et ceux qui ont la permission de validation peuvent valider sans restriction horaire
        # Les utilisateurs normaux sont limités à 13h00 GMT
        # Si l'utilisateur a la permission de validation (has_permission est True ici), il peut valider après 13h00
        peut_valider_apres_limite = est_admin_ou_staff or has_permission
        
        if not peut_valider_apres_limite:
            erreurs_fenetre = []
            for ligne in commande.lignes.all():
                categorie_restau = ligne.menu_plat.plat.categorie_restau
                if not FenetreCommande.est_dans_fenetre(categorie_restau, commande.date_commande, est_public=False):
                    fenetre = FenetreCommande.objects.filter(categorie_restau=categorie_restau, actif=True).first()
                    if fenetre:
                        erreurs_fenetre.append(
                            f"Fenêtre de commande fermée pour {ligne.menu_plat.plat.nom} "
                            f"({ligne.menu_plat.plat.get_categorie_restau_display()}). "
                            f"Heure limite: {fenetre.heure_limite.strftime('%H:%M')}"
                        )
                    else:
                        # Si pas de fenêtre configurée, appliquer la limite par défaut de 13h00 GMT
                        from datetime import time
                        heure_limite_public = time(18, 0, 0)
                        heure_actuelle = timezone.now().time()
                        if commande.date_commande == timezone.now().date() and heure_actuelle > heure_limite_public:
                            erreurs_fenetre.append(
                                f"Fenêtre de commande fermée pour {ligne.menu_plat.plat.nom}. "
                                f"Heure limite: 18:00 GMT"
                            )
            
            if erreurs_fenetre:
                return Response({
                    'error': 'Impossible de valider la commande: fenêtre de commande fermée',
                    'details': erreurs_fenetre
                }, status=400)
        
        try:
            with transaction.atomic():
                # Recalculer les montants
                commande.calculer_montants()
                
                # Créer l'opération budgétaire
                categorie_restauration = Categorie.objects.filter(code='RESTAURATION').first()
                if not categorie_restauration:
                    # Créer la catégorie si elle n'existe pas
                    categorie_restauration = Categorie.objects.create(
                        nom='Restauration',
                        code='RESTAURATION',
                        description='Dépenses de restauration/cantine'
                    )
                
                # Déterminer la sous-catégorie selon le créneau (à améliorer)
                sous_categorie = None
                # Pour l'instant, on utilise la première sous-catégorie de restauration ou on en crée une
                
                # Calculer le nombre total de plats
                nb_plats = sum(ligne.quantite for ligne in commande.lignes.all())
                
                # Créer l'opération
                operation = Operation.objects.create(
                    date_operation=commande.date_commande,
                    categorie=categorie_restauration,
                    sous_categorie=sous_categorie,
                    unites=nb_plats,
                    prix_unitaire=commande.montant_net / nb_plats if nb_plats > 0 else Decimal('0.00'),
                    description=f"Commande restauration - {commande.utilisateur.username}",
                    created_by=request.user
                )
                
                # Lier l'opération à la commande
                commande.operation = operation
                commande.etat = 'validee'
                commande.save()
                
                # Créer l'imputation automatique
                commande.operation.create_imputation_if_needed()
                
                # Générer automatiquement la facture pour ce jour
                try:
                    generer_facture_journaliere(commande.date_commande)
                except Exception as e:
                    print(f"Erreur lors de la génération de la facture: {e}")
                
                # Envoyer un email de confirmation si l'utilisateur a un email
                try:
                    result = envoyer_email_confirmation(commande)
                    if result and result.get('error'):
                        print(f"⚠️ Erreur lors de l'envoi de l'email: {result.get('error')}")
                    elif result and result.get('success'):
                        print(f"✅ Email envoyé avec succès à {result.get('email')}")
                except Exception as e:
                    import traceback
                    print(f"❌ Erreur lors de l'envoi de l'email de confirmation: {e}")
                    print(traceback.format_exc())
                    # Ne pas bloquer la validation si l'email échoue
                
                if request.user.is_authenticated:
                    log_audit('update', request.user, commande, metadata={'action': 'validation_commande', 'operation_id': operation.id})
                
                serializer = self.get_serializer(commande)
                return Response(serializer.data)
        
        except Exception as e:
            return Response({'error': f'Erreur lors de la validation: {str(e)}'}, status=500)
    
    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        """Annuler une commande"""
        commande = self.get_object()
        
        if commande.etat == 'livree':
            return Response({'error': 'Impossible d\'annuler une commande livrée'}, status=400)
        
        commande.etat = 'annulee'
        commande.save()
        
        if request.user.is_authenticated:
            log_audit('update', request.user, commande, metadata={'action': 'annulation_commande'})
        
        serializer = self.get_serializer(commande)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def rapport(self, request):
        """Rapport des commandes pour un mois"""
        mois = request.query_params.get('mois')
        if not mois:
            return Response({'error': 'Paramètre mois requis (format: YYYY-MM)'}, status=400)
        
        try:
            mois_date = datetime.strptime(mois, '%Y-%m').date()
            mois_debut = mois_date.replace(day=1)
            mois_fin = (mois_debut + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        except ValueError:
            return Response({'error': 'Format de date invalide'}, status=400)
        
        commandes = self.queryset.filter(
            date_commande__gte=mois_debut,
            date_commande__lte=mois_fin,
            etat='validee'
        )
        
        total_brut = sum(c.montant_brut for c in commandes)
        total_subvention = sum(c.montant_subvention for c in commandes)
        total_net = sum(c.montant_net for c in commandes)
        nb_commandes = commandes.count()
        nb_plats = sum(sum(l.quantite for l in c.lignes.all()) for c in commandes)
        
        # Top plats
        from django.db.models import Sum
        top_plats = CommandeLigne.objects.filter(
            commande__in=commandes
        ).values(
            'menu_plat__plat__nom'
        ).annotate(
            total_quantite=Sum('quantite')
        ).order_by('-total_quantite')[:10]
        
        return Response({
            'mois': mois,
            'nb_commandes': nb_commandes,
            'nb_plats': nb_plats,
            'total_brut': total_brut,
            'total_subvention': total_subvention,
            'total_net': total_net,
            'moyenne_journaliere': total_net / mois_fin.day if mois_fin.day > 0 else 0,
            'top_plats': list(top_plats)
        })


class CommandeLigneViewSet(viewsets.ModelViewSet):
    """Gestion des lignes de commande"""
    queryset = CommandeLigne.objects.all()
    serializer_class = CommandeLigneSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['commande', 'menu_plat']


class ExtraRestaurationViewSet(viewsets.ModelViewSet):
    """Gestion des extras de restauration (visiteurs, stagiaires, activités)"""
    queryset = ExtraRestauration.objects.select_related('operation', 'created_by').all()
    serializer_class = ExtraRestaurationSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['type_extra', 'date_operation']
    ordering_fields = ['date_operation', 'created_at']
    ordering = ['-date_operation', '-created_at']
    
    def get_queryset(self):
        """Appliquer les filtres de date personnalisés"""
        queryset = super().get_queryset()
        
        # Filtre date_operation_after
        date_after = self.request.query_params.get('date_operation_after')
        if date_after:
            queryset = queryset.filter(date_operation__gte=date_after)
        
        # Filtre date_operation_before
        date_before = self.request.query_params.get('date_operation_before')
        if date_before:
            queryset = queryset.filter(date_operation__lte=date_before)
        
        return queryset
    
    def get_serializer_context(self):
        """Ajouter le contexte (request) au serializer"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_create(self, serializer):
        """Créer l'extra et l'opération associée automatiquement"""
        # Le serializer.create() gère déjà la création de l'opération
        # On passe created_by dans le contexte pour que le serializer puisse l'utiliser
        serializer.save(created_by=self.request.user)


# ============================================================================
# VUES PUBLIQUES (SANS AUTHENTIFICATION)
# ============================================================================

@api_view(['GET'])
@permission_classes([AllowAny])
def menu_public(request, token):
    """Récupérer un menu publié via son token public ou le menu du jour si token='aujourdhui'"""
    from datetime import date
    
    # Si token = "aujourdhui", chercher le menu du jour
    if token == 'aujourdhui':
        try:
            menu = Menu.objects.get(
                date_menu=date.today(),
                publication_at__isnull=False
            )
            serializer = MenuSerializer(menu, context={'request': request})
            return Response(serializer.data)
        except Menu.DoesNotExist:
            return Response({'error': 'Aucun menu publié pour aujourd\'hui'}, status=404)
    
    # Sinon, chercher par token
    try:
        menu = Menu.objects.get(token_public=token, publication_at__isnull=False)
        serializer = MenuSerializer(menu, context={'request': request})
        return Response(serializer.data)
    except Menu.DoesNotExist:
        return Response({'error': 'Menu non trouvé ou non publié'}, status=404)


@api_view(['POST'])
@permission_classes([AllowAny])
def commander_public(request, token):
    """Créer une commande publique via le token du menu ou le menu du jour si token='aujourdhui'
    
    Les commandes publiques sont limitées à 13h00 GMT (heure de Conakry)
    """
    from datetime import date, time
    from django.utils import timezone
    
    # Si token = "aujourdhui", chercher le menu du jour
    if token == 'aujourdhui':
        try:
            menu = Menu.objects.get(
                date_menu=date.today(),
                publication_at__isnull=False
            )
        except Menu.DoesNotExist:
            return Response({'error': 'Aucun menu publié pour aujourd\'hui'}, status=404)
    else:
        # Sinon, chercher par token
        try:
            menu = Menu.objects.get(token_public=token, publication_at__isnull=False)
        except Menu.DoesNotExist:
            return Response({'error': 'Menu non trouvé ou non publié'}, status=404)
    
    # Vérifier la restriction horaire pour les commandes publiques (12h30 GMT)
    heure_limite_public = time(12, 30, 0)
    heure_actuelle = timezone.now().time()
    
    # Si la date de commande est aujourd'hui, vérifier l'heure
    if menu.date_menu == timezone.now().date():
        if heure_actuelle > heure_limite_public:
            return Response({
                'error': 'Les commandes publiques sont fermees',
                'message': 'Les commandes publiques ne sont acceptees que jusqu\'a 12h30 GMT (heure de Conakry)',
                'heure_limite': '12:30',
                'heure_actuelle': heure_actuelle.strftime('%H:%M:%S')
            }, status=403)
    
    # Récupérer les données de la commande
    nom_employe = request.data.get('nom_employe', '')
    email_employe = request.data.get('email_employe', '')
    lignes_data = request.data.get('lignes', [])
    
    if not nom_employe:
        return Response({'error': 'Le nom de l\'employé est requis'}, status=400)
    
    if not lignes_data:
        return Response({'error': 'Aucune ligne de commande'}, status=400)
    
    try:
        with transaction.atomic():
            # Créer un utilisateur unique pour chaque commande (basé sur nom + email + date + timestamp)
            from django.contrib.auth.models import User
            import hashlib
            import time
            # Créer un identifiant unique basé sur le nom, email, date et timestamp
            timestamp = int(time.time() * 1000)  # millisecondes pour plus d'unicité
            identifiant = hashlib.md5(f"{nom_employe}_{email_employe}_{menu.date_menu}_{timestamp}".encode()).hexdigest()[:12]
            username_unique = f'commande_{menu.date_menu.strftime("%Y%m%d")}_{identifiant}'
            
            # Stocker l'email réel si fourni, sinon utiliser un email généré
            email_final = email_employe.strip() if email_employe and email_employe.strip() else f'{identifiant}@commande.local'
            
            user_anonyme = User.objects.create(
                username=username_unique,
                email=email_final,
                first_name=nom_employe,
                is_active=True
            )
            
            # Créer la commande (plusieurs personnes peuvent commander le même jour)
            # Les commandes publiques sont en brouillon, le gestionnaire doit valider
            commande = Commande.objects.create(
                utilisateur=user_anonyme,
                date_commande=menu.date_menu,
                etat='brouillon'  # En attente de validation par le gestionnaire
            )
            
            # Variables pour calculer le supplément
            # Subvention de 30000 GNF uniquement sur le 1er plat (1 seule unite)
            # Les autres plats sont au prix complet
            total_supplement = Decimal('0.00')
            total_subvention = Decimal('0.00')
            plats_avec_supplement = []
            subvention_utilisee = False
            
            # Créer les lignes
            for ligne_data in lignes_data:
                menu_plat_id = ligne_data.get('menu_plat_id')
                quantite = int(ligne_data.get('quantite', 1))
                
                try:
                    menu_plat = MenuPlat.objects.get(pk=menu_plat_id, menu=menu)
                    
                    # Vérifier le stock
                    stock_restant = menu_plat.get_stock_restant()
                    if stock_restant is not None and quantite > stock_restant:
                        raise ValueError(f"Stock insuffisant pour {menu_plat.plat.nom}")
                    
                    ligne = CommandeLigne.objects.create(
                        commande=commande,
                        menu_plat=menu_plat,
                        quantite=quantite,
                        prix_unitaire=menu_plat.prix_jour
                    )
                    
                    # Calculer le supplement pour chaque unite
                    for i in range(quantite):
                        if not subvention_utilisee:
                            # Premier plat: subvention de max 30000 GNF
                            subvention = min(menu_plat.prix_jour, Decimal('30000.00'))
                            total_subvention += subvention
                            supplement_unite = menu_plat.prix_jour - subvention
                            total_supplement += supplement_unite
                            subvention_utilisee = True
                        else:
                            # Autres plats: prix complet a payer
                            total_supplement += menu_plat.prix_jour
                    
                    # Ajouter aux plats avec supplement pour l'affichage
                    if menu_plat.prix_jour > 30000 or quantite > 1 or len(lignes_data) > 1:
                        plats_avec_supplement.append({
                            'plat': menu_plat.plat.nom,
                            'prix_reel': float(menu_plat.prix_jour),
                            'quantite': quantite
                        })
                        
                except MenuPlat.DoesNotExist:
                    raise ValueError(f"MenuPlat {menu_plat_id} non trouvé")
            
            # Calculer les montants
            commande.calculer_montants()
            commande.save()
            
            # Générer automatiquement la facture pour ce jour
            try:
                generer_facture_journaliere(commande.date_commande)
            except Exception as e:
                print(f"Erreur lors de la génération de la facture: {e}")
            
            # L'email sera envoyé quand le gestionnaire validera la commande
            # Pas d'envoi automatique à la création
            
            # Préparer la réponse avec les informations sur le supplément
            serializer = CommandeSerializer(commande)
            response_data = serializer.data
            response_data['supplement_info'] = {
                'total_supplement': float(total_supplement),
                'plats_avec_supplement': plats_avec_supplement,
                'message_supplement': f"Vous devez payer un supplément de {total_supplement:,.0f} GNF en espèces" if total_supplement > 0 else None
            }
            response_data['en_attente_validation'] = True
            
            return Response(response_data, status=201)
    
    except ValueError as e:
        return Response({'error': str(e)}, status=400)
    except Exception as e:
        return Response({'error': f'Erreur: {str(e)}'}, status=500)


def generer_facture_journaliere(date_facture):
    """Génère automatiquement une facture pour une date donnée"""
    from datetime import date
    
    # Vérifier si une facture existe déjà
    facture, created = Facture.objects.get_or_create(
        date_facture=date_facture,
        defaults={'numero_facture': Facture.generer_numero(date_facture)}
    )
    
    # Recalculer les totaux
    try:
        facture.calculer_totaux()
    except Exception as e:
        print(f"Erreur lors du calcul des totaux: {e}")
        raise
    
    # Générer le PDF (toujours régénérer pour éviter les anciens PDFs avec balises HTML)
    try:
        # Supprimer l'ancien PDF s'il existe
        if facture.fichier_pdf and os.path.exists(facture.fichier_pdf.path):
            try:
                os.remove(facture.fichier_pdf.path)
            except:
                pass
        
        pdf_path = generer_pdf_facture(facture)
        if pdf_path and os.path.exists(pdf_path):
            # Utiliser le chemin relatif pour le FileField
            relative_path = os.path.relpath(pdf_path, settings.MEDIA_ROOT)
            facture.fichier_pdf.name = relative_path.replace('\\', '/')
            facture.save()
    except Exception as e:
        print(f"Erreur lors de la génération du PDF: {e}")
        # Ne pas bloquer si le PDF ne peut pas être généré, la facture est quand même créée
    
    return facture


def generer_pdf_facture(facture):
    """Génère un PDF pour une facture"""
    from datetime import date
    
    # Créer le répertoire si nécessaire
    media_root = settings.MEDIA_ROOT if hasattr(settings, 'MEDIA_ROOT') else 'media'
    factures_dir = os.path.join(media_root, 'factures')
    os.makedirs(factures_dir, exist_ok=True)
    
    # Nom du fichier
    filename = f'facture_{facture.date_facture.strftime("%Y%m%d")}.pdf'
    filepath = os.path.join(factures_dir, filename)
    
    # Créer le document PDF avec encodage UTF-8
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Configurer les styles pour supporter UTF-8
    for style_name in styles.byName:
        style = styles[style_name]
        if hasattr(style, 'fontName'):
            # S'assurer que les polices supportent UTF-8
            pass  # ReportLab gère UTF-8 par défaut avec Helvetica
    
    # Récupérer les commandes de la date
    commandes = Commande.objects.filter(
        date_commande=facture.date_facture,
        etat='validee'
    ).select_related('utilisateur').prefetch_related('lignes__menu_plat__plat')
    
    # En-tête avec logo et nom de l'entreprise
    # Chercher le logo dans plusieurs emplacements possibles
    # BASE_DIR pointe vers backend/, donc on remonte d'un niveau pour le répertoire racine
    project_root = settings.BASE_DIR.parent
    logo_paths = [
        str(project_root / 'logocsig.png'),
        str(project_root / 'frontend' / 'src' / 'assets' / 'logocsig.png'),
        str(settings.BASE_DIR / 'logocsig.png'),
    ]
    
    logo_path = None
    for path in logo_paths:
        if os.path.exists(path):
            logo_path = path
            break
    
    # Créer un tableau pour l'en-tête avec logo et texte
    if logo_path:
        try:
            logo = Image(logo_path, width=1.5*inch, height=1.5*inch)
            header_data = [[logo, Paragraph("<b>CSIG</b><br/>FACTURE JOURNALIÈRE - RESTAURATION", styles['Title'])]]
            header_table = Table(header_data, colWidths=[2*inch, 4*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(header_table)
        except Exception as e:
            print(f"Erreur lors du chargement du logo: {e}")
            # Si le logo ne peut pas être chargé, utiliser seulement le texte
            company_name = Paragraph("<b>CSIG</b>", styles['Heading1'])
            company_name.hAlign = 'CENTER'
            elements.append(company_name)
            title = Paragraph(f"<b>FACTURE JOURNALIÈRE - RESTAURATION</b>", styles['Title'])
            title.hAlign = 'CENTER'
            elements.append(title)
    else:
        # Si le logo n'existe pas, utiliser seulement le texte
        company_name = Paragraph("<b>CSIG</b>", styles['Heading1'])
        company_name.hAlign = 'CENTER'
        elements.append(company_name)
        title = Paragraph(f"<b>FACTURE JOURNALIÈRE - RESTAURATION</b>", styles['Title'])
        title.hAlign = 'CENTER'
        elements.append(title)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Informations de la facture
    info_data = [
        ['Numéro de facture:', facture.numero_facture],
        ['Date:', facture.date_facture.strftime('%d/%m/%Y')],
        ['Nombre de commandes:', str(facture.total_commandes)],
    ]
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Détails des commandes - Un seul tableau pour toutes les commandes
    elements.append(Paragraph("<b>DÉTAILS DES COMMANDES</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    # Créer un seul tableau avec toutes les commandes (sans colonne supplément)
    all_lignes_data = [['Commande', 'Plat', 'Quantité', 'Prix unitaire', 'Montant']]
    
    # Fonction pour nettoyer et encoder correctement les chaînes
    def clean_text(text):
        """Nettoie et encode correctement le texte pour le PDF"""
        if text is None:
            return ''
        # Convertir en string et nettoyer
        text = str(text)
        # Remplacer les caractères problématiques
        text = text.replace('\x00', '')  # Supprimer les null bytes
        # Encoder en UTF-8 et décoder pour s'assurer que c'est valide
        try:
            text = text.encode('utf-8', errors='ignore').decode('utf-8')
        except:
            text = text.encode('ascii', errors='ignore').decode('ascii')
        return text.strip()
    
    for commande in commandes:
        # Utiliser le nom (first_name) ou username, nettoyé
        nom_utilisateur = ''
        if commande.utilisateur:
            nom_utilisateur = commande.utilisateur.first_name or commande.utilisateur.username or 'Anonyme'
        else:
            nom_utilisateur = 'Anonyme'
        
        nom_utilisateur = clean_text(nom_utilisateur)
        commande_label = f"#{commande.id} - {nom_utilisateur}"
        
        # Total de la commande (limité à 30 000 GNF par plat)
        total_commande = Decimal('0.00')
        
        # Lignes de la commande
        for ligne in commande.lignes.all():
            # Limiter le prix unitaire à 30 000 GNF pour la facture
            prix_unitaire_facture = min(ligne.prix_unitaire, Decimal('30000.00'))
            montant_ligne_facture = prix_unitaire_facture * ligne.quantite
            total_commande += montant_ligne_facture
            
            # Nettoyer le nom du plat
            nom_plat = clean_text(ligne.menu_plat.plat.nom if ligne.menu_plat and ligne.menu_plat.plat else 'Plat inconnu')
            
            all_lignes_data.append([
                clean_text(commande_label),
                nom_plat,
                str(ligne.quantite),
                f"{prix_unitaire_facture:,.0f} GNF",
                f"{montant_ligne_facture:,.0f} GNF"
            ])
        
        # Ajouter la ligne TOTAL pour cette commande
        all_lignes_data.append([
            clean_text(f"TOTAL #{commande.id}"),
            '',
            '',
            '',
            f"{total_commande:,.0f} GNF"
        ])
        # Ligne vide pour séparer les commandes
        all_lignes_data.append(['', '', '', '', ''])
    
    # Créer le tableau unique avec tous les styles (5 colonnes au lieu de 6)
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]
    
    # Identifier les lignes TOTAL et leur appliquer le style gras
    # Parcourir toutes les lignes et identifier celles qui commencent par "TOTAL"
    for row_idx in range(1, len(all_lignes_data)):
        if len(all_lignes_data[row_idx]) > 0 and str(all_lignes_data[row_idx][0]).startswith('TOTAL'):
            table_styles.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            table_styles.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey))
    
    all_cmd_table = Table(all_lignes_data, colWidths=[2*inch, 2.5*inch, 0.7*inch, 1.2*inch, 1.2*inch])
    all_cmd_table.setStyle(TableStyle(table_styles))
    
    elements.append(all_cmd_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Totaux généraux (sans suppléments, limité à 30 000 GNF par plat)
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph("<b>RÉCAPITULATIF</b>", styles['Heading2']))
    
    # Calculer le total facturé (limité à 30 000 GNF par plat)
    total_facture_brut = Decimal('0.00')
    for commande in commandes:
        for ligne in commande.lignes.all():
            prix_unitaire_facture = min(ligne.prix_unitaire, Decimal('30000.00'))
            total_facture_brut += prix_unitaire_facture * ligne.quantite
    
    # La subvention reste la même (basée sur 30 000 GNF max)
    total_facture_net = total_facture_brut - facture.total_subvention
    
    total_data = [
        ['Total Brut (limité à 30 000 GNF/plat):', f"{total_facture_brut:,.0f} GNF"],
        ['Total Subvention:', f"-{facture.total_subvention:,.0f} GNF"],
        ['Total Net:', f"{total_facture_net:,.0f} GNF"],
    ]
    total_table = Table(total_data, colWidths=[3*inch, 3*inch])
    total_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -2), 'Helvetica'),  # Lignes normales
        ('FONTNAME', (1, -1), (1, -1), 'Helvetica-Bold'),  # Dernière ligne en gras
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
    ]))
    elements.append(total_table)
    
    # Pied de page (sans balises HTML)
    elements.append(Spacer(1, 0.3*inch))
    footer_text = f"Facture générée le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    # Utiliser un style simple sans HTML
    footer_style = styles['Normal']
    footer_style.alignment = 1  # Centré
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)
    
    # Construire le PDF
    doc.build(elements)
    
    return filepath


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def generer_facture(request, date_str):
    """Génère ou récupère une facture pour une date donnée"""
    from datetime import datetime
    
    if not date_str:
        return Response({'error': 'Date requise'}, status=400)
    
    if request.method == 'POST':
        # Générer une nouvelle facture
        try:
            date_facture = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return Response({'error': 'Date invalide. Format attendu: YYYY-MM-DD'}, status=400)
        
        try:
            facture = generer_facture_journaliere(date_facture)
            
            return Response({
                'id': facture.id,
                'numero_facture': facture.numero_facture,
                'date_facture': facture.date_facture.strftime('%Y-%m-%d'),
                'total_commandes': facture.total_commandes,
                'total_brut': str(facture.total_brut),
                'total_subvention': str(facture.total_subvention),
                'total_net': str(facture.total_net),
                'total_supplement': str(facture.total_supplement),
                'fichier_pdf': request.build_absolute_uri(facture.fichier_pdf.url) if facture.fichier_pdf else None,
            })
        except Exception as e:
            return Response({'error': f'Erreur lors de la génération: {str(e)}'}, status=500)
    
    # GET: Récupérer une facture existante
    try:
        date_facture = datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return Response({'error': 'Date invalide. Format attendu: YYYY-MM-DD'}, status=400)
    
    try:
        facture = Facture.objects.get(date_facture=date_facture)
        return Response({
            'id': facture.id,
            'numero_facture': facture.numero_facture,
            'date_facture': facture.date_facture.strftime('%Y-%m-%d'),
            'total_commandes': facture.total_commandes,
            'total_brut': str(facture.total_brut),
            'total_subvention': str(facture.total_subvention),
            'total_net': str(facture.total_net),
            'total_supplement': str(facture.total_supplement),
            'fichier_pdf': request.build_absolute_uri(facture.fichier_pdf.url) if facture.fichier_pdf else None,
        })
    except Facture.DoesNotExist:
        return Response({'error': 'Facture non trouvée pour cette date'}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def imprimer_facture(request, date_str):
    """Télécharge le PDF d'une facture"""
    from datetime import datetime
    
    try:
        date_facture = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return Response({'error': 'Date invalide. Format attendu: YYYY-MM-DD'}, status=400)
    
    try:
        # Récupérer ou créer la facture
        facture, created = Facture.objects.get_or_create(
            date_facture=date_facture,
            defaults={'numero_facture': Facture.generer_numero(date_facture)}
        )
        
        # Recalculer les totaux (sauvegarde automatique dans calculer_totaux)
        facture.calculer_totaux()
        
        # Toujours régénérer le PDF pour éviter les anciens PDFs avec des balises HTML
        pdf_path = None
        try:
            # Supprimer l'ancien PDF s'il existe
            if facture.fichier_pdf and os.path.exists(facture.fichier_pdf.path):
                try:
                    os.remove(facture.fichier_pdf.path)
                except:
                    pass
            
            # Générer un nouveau PDF
            pdf_path = generer_pdf_facture(facture)
            if pdf_path and os.path.exists(pdf_path):
                # Sauvegarder le chemin dans le modèle
                relative_path = os.path.relpath(pdf_path, settings.MEDIA_ROOT)
                facture.fichier_pdf.name = relative_path.replace('\\', '/')
                facture.save()
        except Exception as e:
            print(f"Erreur lors de la génération du PDF: {e}")
            return Response({'error': f'Impossible de générer le PDF: {str(e)}'}, status=500)
        
        if pdf_path and os.path.exists(pdf_path):
            return FileResponse(
                open(pdf_path, 'rb'),
                content_type='application/pdf',
                as_attachment=True,
                filename=f'facture_{facture.date_facture.strftime("%Y%m%d")}.pdf'
            )
        else:
            return Response({'error': 'Fichier PDF introuvable'}, status=404)
            
    except Facture.DoesNotExist:
        return Response({'error': 'Facture non trouvée pour cette date'}, status=404)
    except Exception as e:
        return Response({'error': f'Erreur: {str(e)}'}, status=500)


def envoyer_email_confirmation(commande):
    """Envoie un email de confirmation lorsqu'une commande est validée"""
    print(f"[EMAIL] Tentative d'envoi d'email pour la commande #{commande.id}")
    
    # Récupérer l'email de l'utilisateur
    email_destinataire = None
    nom_employe = None
    
    if commande.utilisateur:
        email_destinataire = commande.utilisateur.email
        # Pour les commandes publiques, le nom est dans first_name
        nom_employe = commande.utilisateur.first_name or commande.utilisateur.username
        print(f"   Utilisateur: {commande.utilisateur.username}")
        print(f"   Email trouve: {email_destinataire}")
        
        # Si l'email est un email généré (commande publique), ne pas envoyer
        if email_destinataire and '@commande.local' in email_destinataire:
            print(f"   [INFO] Email non fourni par l'utilisateur, pas d'envoi")
            return {'success': False, 'error': 'Email non fourni', 'email': None}
    else:
        print(f"   [INFO] Aucun utilisateur associe a la commande")
        return {'success': False, 'error': 'Aucun utilisateur associé à la commande', 'email': None}
    
    # Si pas d'email valide, ne pas envoyer
    if not email_destinataire or not email_destinataire.strip() or '@' not in email_destinataire:
        print(f"   [INFO] Email invalide ou manquant: {email_destinataire}")
        return {'success': False, 'error': 'Aucun email valide trouve', 'email': email_destinataire or 'N/A'}
    
    print(f"   [OK] Email valide: {email_destinataire}")
    
    # Préparer les données pour le template
    lignes = []
    total_supplement = 0
    total_plats_simple = 0  # Total des plats <= 30 000 GNF
    
    for ligne in commande.lignes.all():
        prix_unitaire = float(ligne.prix_unitaire)
        
        if prix_unitaire > 30000:
            # Pour les plats > 30 000 GNF, on calcule seulement le supplément
            supplement_ligne = (prix_unitaire - 30000) * ligne.quantite
            total_supplement += supplement_ligne
            lignes.append({
                'plat_nom': ligne.menu_plat.plat.nom,
                'quantite': ligne.quantite,
                'prix_unitaire': f"{prix_unitaire:,.0f}",
                'montant_ligne': f"{ligne.montant_ligne:,.0f}",
                'supplement': f"{supplement_ligne:,.0f}",
                'depasse_30000': True,
            })
        else:
            # Pour les plats <= 30 000 GNF, on affiche le montant normal
            total_plats_simple += float(ligne.montant_ligne)
            lignes.append({
                'plat_nom': ligne.menu_plat.plat.nom,
                'quantite': ligne.quantite,
                'prix_unitaire': f"{prix_unitaire:,.0f}",
                'montant_ligne': f"{ligne.montant_ligne:,.0f}",
                'supplement': 0,
                'depasse_30000': False,
            })
    
    # Calculer le montant net à payer selon les règles
    # Si seulement des plats simples (<= 30 000), montant_net = total_plats_simple
    # Si seulement des plats avec supplément, montant_net = total_supplement
    # Si mixte, montant_net = total_plats_simple + total_supplement
    montant_net_a_payer = total_plats_simple + total_supplement
    
    context = {
        'nom_employe': nom_employe or commande.utilisateur.username if commande.utilisateur else 'Client',
        'date_commande': commande.date_commande.strftime('%d/%m/%Y'),
        'commande_id': commande.id,
        'lignes': lignes,
        'montant_net_a_payer': f"{montant_net_a_payer:,.0f}",
        'total_plats_simple': f"{total_plats_simple:,.0f}",
        'total_supplement': f"{total_supplement:,.0f}",
        'a_supplement': total_supplement > 0,
        'a_plats_simples': total_plats_simple > 0,
    }
    
    # Rendre le template HTML
    html_message = render_to_string('emails/commande_validee.html', context)
    plain_message = strip_tags(html_message)
    
    # Envoyer l'email
    from django.conf import settings
    print(f"   Configuration SMTP:")
    print(f"      Host: {settings.EMAIL_HOST}")
    print(f"      Port: {settings.EMAIL_PORT}")
    print(f"      TLS: {settings.EMAIL_USE_TLS}")
    print(f"      User: {settings.EMAIL_HOST_USER}")
    print(f"      From: {settings.DEFAULT_FROM_EMAIL or 'support@csig.edu.gn'}")
    
    try:
        print(f"   [ENVOI] Envoi de l'email en cours...")
        send_mail(
            subject=f'Confirmation de commande #{commande.id} - CSIG',
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL or 'support@csig.edu.gn',
            recipient_list=[email_destinataire],
            html_message=html_message,
            fail_silently=False,
        )
        print(f"   [OK] Email envoye avec succes a {email_destinataire}")
        return {'success': True, 'email': email_destinataire}
    except Exception as e:
        import traceback
        error_msg = f"Erreur SMTP: {str(e)}"
        print(f"   [ERREUR] {error_msg}")
        print(f"   Details:")
        print(traceback.format_exc())
        return {'success': False, 'error': error_msg, 'email': email_destinataire}

