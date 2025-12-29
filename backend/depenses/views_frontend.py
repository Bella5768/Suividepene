from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate
from django.conf import settings
from django.template.loader import get_template
from rest_framework_simplejwt.tokens import RefreshToken
import json
import logging
import calendar
from datetime import datetime
from decimal import Decimal
from django.db.models import Sum, Count

logger = logging.getLogger(__name__)


def index(request, **kwargs):
    """Vue principale qui sert l'interface frontend vanilla
    
    Accepte des paramètres supplémentaires (comme token) pour les routes
    qui nécessitent des paramètres dans l'URL, mais le routeur vanilla JS gère
    le routage côté client.
    """
    try:
        template = get_template('depenses/index.html')
        context = {
            'user': request.user,
            'debug': settings.DEBUG,
        }
        return HttpResponse(template.render(context, request))
    except Exception as e:
        logger.error(f"Erreur lors du chargement du template: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)


def get_token(request):
    """Endpoint pour obtenir un token JWT (pour l'API React)"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')
            
            user = authenticate(request, username=username, password=password)
            if user:
                refresh = RefreshToken.for_user(user)
                return JsonResponse({
                    'access': str(refresh.access_token),
                    'refresh': str(refresh),
                })
            else:
                return JsonResponse({'error': 'Invalid credentials'}, status=401)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def export_rapport_pdf(request):
    """Exporter le rapport mensuel en PDF - Vue Django avec login_required"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from .models import Operation, Prevision
    
    mois = request.GET.get('mois')
    if not mois:
        return HttpResponse('Paramètre mois requis', status=400)
    
    try:
        mois_date = datetime.strptime(mois, '%Y-%m').date()
        mois_date = mois_date.replace(day=1)
    except ValueError:
        return HttpResponse('Format de date invalide', status=400)
    
    # Calculer le dernier jour du mois
    dernier_jour = calendar.monthrange(mois_date.year, mois_date.month)[1]
    date_fin = mois_date.replace(day=dernier_jour)
    
    # Opérations du mois
    operations = Operation.objects.filter(
        date_operation__gte=mois_date,
        date_operation__lte=date_fin
    )
    
    # Prévisions du mois
    previsions = Prevision.objects.filter(mois=mois_date)
    
    # Totaux
    total_depenses = operations.aggregate(Sum('montant_depense'))['montant_depense__sum'] or 0
    total_prevu = previsions.aggregate(Sum('montant_prevu'))['montant_prevu__sum'] or 0
    nombre_operations = operations.count()
    
    # Moyenne journalière
    jours_avec_operations = operations.values('date_operation').distinct().count()
    moyenne_journaliere = total_depenses / jours_avec_operations if jours_avec_operations > 0 else 0
    
    # Totaux par catégorie
    totals_operations = operations.values('categorie__code', 'categorie__nom').annotate(
        total=Sum('montant_depense'),
        count=Count('id')
    )
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapport_{mois}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    title = Paragraph(f"Rapport Mensuel - {mois}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Informations générales
    info_data = [
        ['Période', f"{mois_date} au {date_fin}"],
        ['Total Dépenses', f"{total_depenses:,.2f} GNF"],
        ['Total Prévu', f"{total_prevu:,.2f} GNF"],
        ['Écart Global', f"{total_depenses - total_prevu:,.2f} GNF"],
        ['Nombre d\'opérations', str(nombre_operations)],
        ['Moyenne Journalière', f"{moyenne_journaliere:,.2f} GNF"],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Tableau des catégories
    cat_headers = ['Catégorie', 'Total Dépense', 'Montant Prévu', 'Écart', 'Nb Opérations']
    cat_data = [cat_headers]
    
    for tot in totals_operations:
        categorie_code = tot['categorie__code']
        prevision_cat = previsions.filter(categorie__code=categorie_code).first()
        montant_prevu = prevision_cat.montant_prevu if prevision_cat else 0
        ecart = tot['total'] - montant_prevu
        
        cat_data.append([
            f"{categorie_code} - {tot['categorie__nom']}",
            f"{tot['total']:,.2f} GNF",
            f"{montant_prevu:,.2f} GNF",
            f"{ecart:,.2f} GNF",
            str(tot['count'])
        ])
    
    if len(cat_data) > 1:
        cat_table = Table(cat_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(cat_table)
    
    doc.build(elements)
    return response


@login_required
def export_rapport_excel(request):
    """Exporter le rapport mensuel en Excel - Vue Django avec login_required"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from .models import Operation, Prevision
    
    mois = request.GET.get('mois')
    if not mois:
        return HttpResponse('Paramètre mois requis', status=400)
    
    try:
        mois_date = datetime.strptime(mois, '%Y-%m').date()
        mois_date = mois_date.replace(day=1)
    except ValueError:
        return HttpResponse('Format de date invalide', status=400)
    
    # Calculer le dernier jour du mois
    dernier_jour = calendar.monthrange(mois_date.year, mois_date.month)[1]
    date_fin = mois_date.replace(day=dernier_jour)
    
    # Opérations du mois
    operations = Operation.objects.filter(
        date_operation__gte=mois_date,
        date_operation__lte=date_fin
    )
    
    # Prévisions du mois
    previsions = Prevision.objects.filter(mois=mois_date)
    
    # Totaux
    total_depenses = operations.aggregate(Sum('montant_depense'))['montant_depense__sum'] or 0
    total_prevu = previsions.aggregate(Sum('montant_prevu'))['montant_prevu__sum'] or 0
    nombre_operations = operations.count()
    
    # Moyenne journalière
    jours_avec_operations = operations.values('date_operation').distinct().count()
    moyenne_journaliere = total_depenses / jours_avec_operations if jours_avec_operations > 0 else 0
    
    # Totaux par catégorie
    totals_operations = operations.values('categorie__code', 'categorie__nom').annotate(
        total=Sum('montant_depense'),
        count=Count('id')
    )
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rapport {mois}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    
    # Titre
    ws['A1'] = f"Rapport Mensuel - {mois}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    # Informations générales
    ws['A3'] = "Période"
    ws['B3'] = f"{mois_date} au {date_fin}"
    ws['A4'] = "Total Dépenses"
    ws['B4'] = float(total_depenses)
    ws['A5'] = "Total Prévu"
    ws['B5'] = float(total_prevu)
    ws['A6'] = "Écart Global"
    ws['B6'] = float(total_depenses - total_prevu)
    ws['A7'] = "Nombre d'opérations"
    ws['B7'] = nombre_operations
    ws['A8'] = "Moyenne Journalière"
    ws['B8'] = float(moyenne_journaliere)
    
    # En-têtes du tableau des catégories
    headers = ['Catégorie', 'Total Dépense', 'Montant Prévu', 'Écart', 'Nb Opérations']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Données par catégorie
    row = 11
    for tot in totals_operations:
        categorie_code = tot['categorie__code']
        prevision_cat = previsions.filter(categorie__code=categorie_code).first()
        montant_prevu = float(prevision_cat.montant_prevu) if prevision_cat else 0
        ecart = float(tot['total']) - montant_prevu
        
        ws.cell(row=row, column=1, value=f"{categorie_code} - {tot['categorie__nom']}")
        ws.cell(row=row, column=2, value=float(tot['total']))
        ws.cell(row=row, column=3, value=montant_prevu)
        ws.cell(row=row, column=4, value=ecart)
        ws.cell(row=row, column=5, value=tot['count'])
        row += 1
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # Créer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="rapport_{mois}.xlsx"'
    wb.save(response)
    return response

